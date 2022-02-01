import os
from decouple import config
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

def make_env():
    output_path = config('OUTPUTPATH')
    if not os.path.exists(output_path):
        os.mkdir(output_path)

def main():

    flag = True
    suma = 0
    counter = 0
    headers = []
    columns = []

    # Apertura del template
    template = load_workbook('template.xlsx')
    sheet = template.active

    # Row donde empezamos a pegar los datos
    row_inicial = 3

    # Instanciamos el driver para comunicarnos con el navegador
    driver = webdriver.Chrome()
    driver.get(config('URL'))

    # La tabla tiene como id billetes
    elemento = driver.find_element(By.ID, "billetes")

    # Cada una de las filas esta dentro de un elemento tr
    rows = elemento.find_elements(By.TAG_NAME, "tr")

    # Iteramos por cada elemento tr encontrado (filas)
    for row in rows:

        # La primera vez debemos extraer los headers donde se encuentra la fecha
        if flag:
            flag = False
            headers = row.find_elements(By.TAG_NAME, "th")

        columns = row.find_elements(By.TAG_NAME, "td")
        # Debido a que la fecha tambien esta dentro de un elemento tr, pero no contiene 
        # elementos td, la primera vez columns sera una lista vacia
        if columns != []:
            for c in range(len(columns)):
                value = columns[c].text.replace(',','.')
                # Intentamos castear el valor a float
                try:
                    value = float(value)
                except:
                    pass
                sheet.cell(row=row_inicial, column=c+1).value = value
                if counter != 0:
                    suma = suma + float(value)
                counter = counter + 1
            sheet.cell(row=row_inicial, column=counter+1).value = suma / 2 
            row_inicial = row_inicial + 1
            counter = 0
            suma = 0
    
    fecha = headers[0].text

    sheet.cell(row=1, column=2).value = fecha

    fecha = fecha.replace('/','-')

    output_file = "cotizaciones {}.xlsx".format(fecha)

    template.save(config('OUTPUTPATH') + output_file)

if __name__ == '__main__':
    main()